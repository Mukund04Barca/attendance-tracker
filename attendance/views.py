from datetime import date, datetime, timedelta
import logging
import json
import calendar
from io import BytesIO

import pandas as pd
import yaml
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.core.mail import send_mail
from django.utils import timezone
from django_ratelimit.decorators import ratelimit

from .models import AttendanceRecord, Holiday, TimesheetActivity, CompOffRecord, TimesheetRecord

logger = logging.getLogger("attendance")

ATT_CFG = getattr(settings, "ATTENDANCE_CONFIG", {})
MAX_TIMESHEET_ACTIVITIES = 100
MAX_TEXT_FIELD_LENGTH = 200

WORKDAY_START = datetime.strptime(
    ATT_CFG.get("workday_start", "09:00"), "%H:%M"
).time()
WORKDAY_END = datetime.strptime(
    ATT_CFG.get("workday_end", "18:00"), "%H:%M"
).time()
DEFAULT_ALLOWANCE = float(ATT_CFG.get("default_allowance_hours", 0.0))
CONFIG_HOLIDAYS = set(ATT_CFG.get("holidays", []))
WEEKLY_TARGET = float(ATT_CFG.get("weekly_hours_target", 45))
DAILY_TARGET_HOURS = float(ATT_CFG.get("daily_hours_target", 9))
SATURDAY_TARGET_HOURS = float(ATT_CFG.get("saturday_hours_target", 6))


def _clean_text(value, limit=MAX_TEXT_FIELD_LENGTH):
    return str(value or "").strip()[:limit]


def _clean_timesheet_activities(raw_activities, num_days):
    if not isinstance(raw_activities, list):
        raise ValueError("Timesheet payload must be a list.")
    if len(raw_activities) > MAX_TIMESHEET_ACTIVITIES:
        raise ValueError("Too many timesheet rows submitted.")

    cleaned = []
    for i, act_data in enumerate(raw_activities):
        if not isinstance(act_data, dict):
            raise ValueError("Invalid timesheet row.")

        raw_hours = act_data.get("hours", {})
        if not isinstance(raw_hours, dict):
            raise ValueError("Invalid hours payload.")

        hours = {}
        for day_key, raw_value in raw_hours.items():
            day = int(day_key)
            if day < 1 or day > num_days:
                continue
            if raw_value in ("", None):
                continue

            value = float(raw_value)
            if value < 0 or value > 24:
                raise ValueError("Daily hours must be between 0 and 24.")
            hours[str(day)] = value

        cleaned.append(
            {
                "sr": i + 1,
                "category": _clean_text(act_data.get("category")),
                "sub": _clean_text(act_data.get("sub"), 500),
                "id": _clean_text(act_data.get("id"), 100),
                "hours": hours,
            }
        )

    return cleaned


def landing_view(request):
    if request.user.is_authenticated:
        return redirect("checkin_checkout")
    return render(request, "attendance/landing.html")


def is_config_holiday(date):
    return date.isoformat() in CONFIG_HOLIDAYS


def build_billable_time_entry(record):
    """Normalize timesheet display/export to the standard billing window."""
    if record.leave_type or record.is_holiday:
        return {
            "in_time": "LEAVE" if record.leave_type else "HOLIDAY",
            "out_time": "—",
            "total_time": "00:00",
            "esa_time": 0,
        }

    # Check for derived timesheet data first
    try:
        derived = record.timesheet_derived
        ci_dt = derived.ts_check_in
        co_dt = derived.ts_check_out
    except:
        ci_dt = record.check_in
        co_dt = None # Will calculate below

    if not ci_dt:
        return None

    ci_local = timezone.localtime(ci_dt)
    if co_dt:
        co_local = timezone.localtime(co_dt)
    else:
        # Fallback calculation if no specific derived checkout exists
        co_local = ci_local + timedelta(hours=DAILY_TARGET_HOURS)
    
    billable_delta = co_local - ci_local
    total_sec = max(billable_delta.total_seconds(), 0)
    hh = int(total_sec // 3600)
    mm = int((total_sec % 3600) // 60)

    return {
        "in_time": ci_local.strftime("%H:%M"),
        "out_time": co_local.strftime("%H:%M"),
        "total_time": f"{hh:02d}:{mm:02d}",
        "esa_time": round(total_sec / 3600.0, 2),
    }

@ratelimit(key='ip', rate='100/m', block=True)
@login_required
def checkin_checkout_view(request):
    now = timezone.localtime(timezone.now())
    today = now.date()

    record, created = AttendanceRecord.objects.get_or_create(
        user=request.user,
        date=today,
        defaults={
            "is_holiday": Holiday.objects.filter(date=today).exists()
            or is_config_holiday(today),
            "allowance_hours": DEFAULT_ALLOWANCE if is_config_holiday(today) else 0.0,
        },
    )

    if request.method == "POST":
        action = request.POST.get("action")

        if action in {"check_in", "checkin"} and record.check_in is None:
            record.check_in = now
            record.save()
            logger.info(
                "User %s checked in at %s", request.user.username, now.isoformat()
            )
        elif action in {"check_out", "checkout"} and record.check_out is None and record.check_in:
            record.check_out = now
            record.save()
            logger.info(
                "User %s checked out at %s", request.user.username, now.isoformat()
            )
        elif action == "set_manual":
            check_in_str = request.POST.get("manual_check_in") or ""
            check_out_str = request.POST.get("manual_check_out") or ""
            try:
                if check_in_str:
                    ci_naive = datetime.strptime(
                        f"{today} {check_in_str}", "%Y-%m-%d %H:%M"
                    )
                    record.check_in = timezone.make_aware(ci_naive)
                if check_out_str:
                    co_naive = datetime.strptime(
                        f"{today} {check_out_str}", "%Y-%m-%d %H:%M"
                    )
                    record.check_out = timezone.make_aware(co_naive)
                record.save()
                logger.info(
                    "User %s set manual times: %s - %s",
                    request.user.username,
                    check_in_str or "—",
                    check_out_str or "—",
                )
            except ValueError:
                pass
        elif action == "delete":
            logger.info(
                "User %s deleted attendance for %s",
                request.user.username,
                today.isoformat(),
            )
            record.delete()
            return redirect("checkin_checkout")

        return redirect("weekly_summary")

    hours_today = None
    if record.check_in and record.check_out:
        delta = record.check_out - record.check_in
        hours_today = round(max(delta.total_seconds() / 3600.0, 0), 2)

    check_in_display = None
    check_out_display = None
    if record.check_in:
        ci_local = timezone.localtime(record.check_in)
        check_in_display = ci_local.strftime("%H:%M")
    if record.check_out:
        co_local = timezone.localtime(record.check_out)
        check_out_display = co_local.strftime("%H:%M")

    expected_checkout = None
    if record.check_in:
        expected_dt = record.check_in + timedelta(hours=DAILY_TARGET_HOURS)
        expected_local = timezone.localtime(expected_dt)
        expected_checkout = expected_local.strftime("%H:%M")

    week_start = today - timedelta(days=today.weekday())
    weekly_records = list(
        AttendanceRecord.objects.filter(
            user=request.user,
            date__range=(week_start, today),
            check_in__isnull=False,
        )
    )
    completed_week_records = [r for r in weekly_records if r.check_out]
    weekly_total = round(
        sum(
            max((r.check_out - r.check_in).total_seconds() / 3600.0, 0)
            for r in completed_week_records
        ),
        1,
    )
    weekly_total_percent = min(int((weekly_total / WEEKLY_TARGET) * 100), 100) if WEEKLY_TARGET else 0
    week_completion_percent = (
        min(int((len(completed_week_records) / len(weekly_records)) * 100), 100)
        if weekly_records
        else 0
    )

    context = {
        "record": record,
        "hours_today": hours_today,
        "check_in_display": check_in_display,
        "check_out_display": check_out_display,
        "expected_checkout": expected_checkout,
        "daily_target_hours": DAILY_TARGET_HOURS,
        "today": today,
        "upcoming_holidays": list(
            Holiday.objects.filter(date__gte=today).order_by("date")[:3]
        ),
        "weekly_total": weekly_total,
        "weekly_target": WEEKLY_TARGET,
        "weekly_total_percent": weekly_total_percent,
        "week_completed_days": len(completed_week_records),
        "week_recorded_days": len(weekly_records),
        "week_completion_percent": week_completion_percent,
        "recent_records": AttendanceRecord.objects.filter(user=request.user, date__lt=today).order_by("-date")[:5],
        "is_leave": bool(record.leave_type),
    }
    return render(request, "attendance/checkin_checkout.html", context)


@login_required
def weekly_summary_view(request):
    today = timezone.localtime(timezone.now()).date()
    start_date = today - timedelta(days=today.weekday())
    end_date = start_date + timedelta(days=6)

    qs = (
        AttendanceRecord.objects.filter(
            user=request.user,
            date__range=(start_date, end_date),
        )
        .order_by("date")
        .values("date", "check_in", "check_out", "is_holiday", "allowance_hours", "leave_type")
    )

    df = pd.DataFrame(list(qs))
    summary = {}

    working_days = 0
    for i in range(7):
        d = start_date + timedelta(days=i)
        if d.weekday() >= 5:
            continue
        if Holiday.objects.filter(date=d).exists() or is_config_holiday(d):
            continue
        working_days += 1
    weekly_target = round(working_days * DAILY_TARGET_HOURS, 2)

    if not df.empty:
        df["check_in"] = pd.to_datetime(df["check_in"], utc=True).dt.tz_convert(
            settings.TIME_ZONE
        ).dt.tz_localize(None)
        df["check_out"] = pd.to_datetime(df["check_out"], utc=True).dt.tz_convert(
            settings.TIME_ZONE
        ).dt.tz_localize(None)

        df["hours"] = (
            (df["check_out"] - df["check_in"])
            .dt.total_seconds()
            .fillna(0)
            / 3600.0
        ).clip(lower=0)

        df["effective_hours"] = df.apply(
            lambda row: 0.0 if row["is_holiday"] else row["hours"], axis=1
        )
        df["total_with_allowance"] = df["effective_hours"] + df["allowance_hours"]
        df["check_in_time"] = df["check_in"].dt.strftime("%H:%M").fillna("")
        df["check_out_time"] = df["check_out"].dt.strftime("%H:%M").fillna("")

        weekly_avg = df["total_with_allowance"].mean()
        weekly_total = df["total_with_allowance"].sum()

        summary = {
            "rows": df.to_dict(orient="records"),
            "weekly_avg": round(float(weekly_avg), 2),
            "weekly_total": round(float(weekly_total), 2),
        }

    context = {
        "summary": summary,
        "start_date": start_date,
        "end_date": end_date,
        "weekly_target": weekly_target,
    }
    return render(request, "attendance/weekly_summary.html", context)

@ratelimit(key='ip', rate='10/m', block=True)
def signup_view(request):
    if request.user.is_authenticated:
        return redirect("timesheet")

    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            logger.info("New user signed up: %s", user.username)
            return redirect("timesheet")
    else:
        form = UserCreationForm()

    return render(request, "registration/signup.html", {"form": form})


@login_required
def month_calendar_view(request):
    year_val = request.GET.get("year")
    month_val = request.GET.get("month")

    today = timezone.localtime(timezone.now()).date()
    if year_val and month_val:
        try:
            year = int(year_val)
            month = int(month_val)
            if month < 1 or month > 12:
                raise ValueError
            current = date(year, month, 1)
        except (TypeError, ValueError):
            current = date(today.year, today.month, 1)
            year, month = today.year, today.month
    else:
        current = date(today.year, today.month, 1)
        year, month = today.year, today.month

    cal = calendar.Calendar(firstweekday=0)
    month_weeks_raw = cal.monthdatescalendar(year, month)

    records = AttendanceRecord.objects.filter(
        user=request.user,
        date__year=year,
        date__month=month,
    )

    records_by_date = {}
    total_extra_hours = 0.0
    total_short_hours = 0.0
    saturday_comp_off_eligible = 0

    def format_duration(decimal_hours, show_sign=False):
        is_neg = decimal_hours < 0
        total_mins = int(round(abs(decimal_hours) * 60))
        h = total_mins // 60
        m = total_mins % 60
        sign = ""
        if show_sign:
            sign = "-" if is_neg else "+"
        if h == 0: return f"{sign}{m}m"
        if m == 0: return f"{sign}{h}h"
        return f"{sign}{h}h {m}m"

    for r in records:
        ci_str = co_str = None
        hours = 0.0
        extra_this_day = 0.0
        if r.check_in:
            ci_str = timezone.localtime(r.check_in).strftime("%H:%M")
        if r.check_out:
            co_str = timezone.localtime(r.check_out).strftime("%H:%M")
        if r.check_in and r.check_out:
            delta = r.check_out - r.check_in
            hours = round(max(delta.total_seconds() / 3600.0, 0), 2)
            if r.date.weekday() < 5:
                if hours > DAILY_TARGET_HOURS:
                    extra_this_day = hours - DAILY_TARGET_HOURS
                    total_extra_hours += extra_this_day
                else:
                    total_short_hours += (DAILY_TARGET_HOURS - hours)
            elif r.date.weekday() == 5:
                if hours >= SATURDAY_TARGET_HOURS:
                    saturday_comp_off_eligible += 1
                    extra_this_day = hours - SATURDAY_TARGET_HOURS
                    total_extra_hours += extra_this_day
            else:
                extra_this_day = hours
                total_extra_hours += extra_this_day

        balance = hours - DAILY_TARGET_HOURS if r.date.weekday() < 5 else hours
        records_by_date[r.date] = {
            "check_in": ci_str,
            "check_out": co_str,
            "hours": hours,
            "hours_display": format_duration(hours),
            "extra_display": format_duration(extra_this_day, show_sign=True) if extra_this_day > 0 else "",
            "balance_display": format_duration(balance, show_sign=True),
            "is_leave": bool(r.leave_type),
        }

    holiday_dates = {h.date for h in Holiday.objects.filter(date__year=year, date__month=month)}
    working_days = 0
    _, days_in_month = calendar.monthrange(year, month)
    for d_idx in range(1, days_in_month + 1):
        d = date(year, month, d_idx)
        if is_config_holiday(d): holiday_dates.add(d)
        if d.weekday() < 5 and d not in holiday_dates: working_days += 1

    prev_month = (current.replace(day=1) - timedelta(days=1)).replace(day=1)
    next_month = (current.replace(day=28) + timedelta(days=4)).replace(day=1)

    weeks = []
    for week in month_weeks_raw:
        row = []
        for d in week:
            row.append({
                "date": d,
                "is_current_month": d.month == month,
                "is_holiday": d in holiday_dates,
                "record": records_by_date.get(d),
            })
        weeks.append(row)

    total_hours_worked = round(sum(r["hours"] for r in records_by_date.values()), 2)
    target_hours = working_days * DAILY_TARGET_HOURS
    extra_hours_in_hand = round(total_extra_hours - total_short_hours, 2)
    
    # Progress Calculation
    progress_percent = min(int((total_hours_worked / target_hours * 100) if target_hours > 0 else 0), 100)
    remaining_hours = max(target_hours - total_hours_worked, 0)

    context = {
        "current": current, "weeks": weeks, "today": today,
        "prev_year": prev_month.year, "prev_month": prev_month.month,
        "next_year": next_month.year, "next_month": next_month.month,
        "monthly_summary": {
            "working_days": working_days,
            "target_hours": target_hours,
            "target_hours_display": format_duration(target_hours),
            "total_hours": total_hours_worked,
            "total_hours_display": format_duration(total_hours_worked),
            "extra_hours_in_hand": extra_hours_in_hand,
            "extra_hours_display": format_duration(extra_hours_in_hand),
            "remaining_to_target_display": format_duration(remaining_hours),
            "bank_progress_percent": progress_percent,
            "comp_off_eligible": saturday_comp_off_eligible,
            "earned_extra_display": format_duration(total_extra_hours),
            "short_hours_display": format_duration(total_short_hours),
        }
    }
    return render(request, "attendance/month_calendar.html", context)


@login_required
def holiday_list_view(request):
    holidays = Holiday.objects.all().order_by("date")
    if request.method == "POST" and request.user.is_staff:
        delete_id = request.POST.get("delete_holiday")
        if delete_id: Holiday.objects.filter(id=delete_id).delete()
        else:
            date_str = request.POST.get("date")
            name = request.POST.get("name") or "Holiday"
            try:
                h_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                Holiday.objects.get_or_create(date=h_date, defaults={"name": name})
            except: pass
        return redirect("holiday_list")
    return render(request, "attendance/holiday_list.html", {"holidays": holidays})


@login_required
def month_excel_export_view(request):
    today = timezone.localtime(timezone.now()).date()
    try:
        year = int(request.GET.get("year", today.year))
        month = int(request.GET.get("month", today.month))
        if month < 1 or month > 12:
            raise ValueError
    except (TypeError, ValueError):
        year, month = today.year, today.month
    qs = AttendanceRecord.objects.filter(user=request.user, date__year=year, date__month=month).values("date", "check_in", "check_out", "is_holiday", "allowance_hours", "leave_type")
    df = pd.DataFrame(list(qs))
    for col in ("check_in", "check_out"):
        if col in df.columns and not df.empty:
            df[col] = pd.to_datetime(df[col], utc=True, errors="coerce").dt.tz_convert(settings.TIME_ZONE).dt.tz_localize(None)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        (df if not df.empty else pd.DataFrame()).to_excel(writer, index=False, sheet_name="Attendance")
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="attendance_{year}_{month}.xlsx"'
    return response


@login_required
def delete_record_view(request, record_date):
    """Delete an AttendanceRecord for any date from the month calendar."""
    if request.method != "POST":
        return redirect("month_calendar")
    try:
        target_date = datetime.strptime(record_date, "%Y-%m-%d").date()
        AttendanceRecord.objects.filter(user=request.user, date=target_date).delete()
        logger.info("User %s deleted record for %s", request.user.username, record_date)
    except: pass
    url = reverse("month_calendar")
    return redirect(f"{url}?year={target_date.year}&month={target_date.month}")


@login_required
def edit_record_view(request, record_date):
    """Create, update, or delete an AttendanceRecord's check-in/check-out for any date."""
    try:
        target_date = datetime.strptime(record_date, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return redirect("month_calendar")

    if request.method != "POST":
        return redirect("month_calendar")

    action = request.POST.get("action")
    if action == "delete":
        AttendanceRecord.objects.filter(user=request.user, date=target_date).delete()
        logger.info("User %s deleted record for %s", request.user.username, record_date)
        url = reverse("month_calendar")
        return redirect(f"{url}?year={target_date.year}&month={target_date.month}")

    check_in_str = request.POST.get("check_in", "").strip()
    check_out_str = request.POST.get("check_out", "").strip()
    mark_leave = action == "leave"

    record, _ = AttendanceRecord.objects.get_or_create(
        user=request.user,
        date=target_date,
        defaults={
            "is_holiday": Holiday.objects.filter(date=target_date).exists() or is_config_holiday(target_date),
            "allowance_hours": DEFAULT_ALLOWANCE if is_config_holiday(target_date) else 0.0,
        },
    )

    try:
        if mark_leave:
            record.check_in = None
            record.check_out = None
            record.leave_type = "Leave"
        else:
            record.leave_type = None
            if check_in_str:
                ci_naive = datetime.strptime(f"{target_date} {check_in_str}", "%Y-%m-%d %H:%M")
                record.check_in = timezone.make_aware(ci_naive)
            else: record.check_in = None
            if check_out_str:
                co_naive = datetime.strptime(f"{target_date} {check_out_str}", "%Y-%m-%d %H:%M")
                record.check_out = timezone.make_aware(co_naive)
            else: record.check_out = None
        record.save()
    except: pass
    url = reverse("month_calendar")
    return redirect(f"{url}?year={target_date.year}&month={target_date.month}")


# ─────────────────────────────────────────────────────────
# TIMESHEET
# ─────────────────────────────────────────────────────────

@login_required
def timesheet_view(request):
    """Display and manage the NSE-format monthly timesheet independently of roster."""
    today = timezone.localtime(timezone.now()).date()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if month < 1 or month > 12:
            raise ValueError
    except (TypeError, ValueError):
        year, month = today.year, today.month

    current = date(year, month, 1)
    num_days = calendar.monthrange(year, month)[1]
    days = list(range(1, num_days + 1))

    def get_billable_entry(record):
        # Try to get derived data from the separate table
        try:
            derived = record.timesheet_derived
            ci_dt = derived.ts_check_in
            co_dt = derived.ts_check_out
        except:
            # Fallback to source only if derivation hasn't happened
            ci_dt = record.check_in
            co_dt = record.check_out

        if not ci_dt:
            if record.leave_type: return {'type': 'LEAVE', 'in': 'OFF', 'out': 'OFF'}
            if record.is_holiday: return {'type': 'HOLIDAY', 'in': 'HOL', 'out': 'HOL'}
            return None
        
        ci_local = timezone.localtime(ci_dt)
        if co_dt:
            co_local = timezone.localtime(co_dt)
        else:
            # Logic: In + 9h for weekdays, In + 6h for Saturdays
            day_target = DAILY_TARGET_HOURS if record.date.weekday() < 5 else SATURDAY_TARGET_HOURS
            co_local = ci_local + timedelta(hours=day_target)
            
        return {'in': ci_local.strftime('%H:%M'), 'out': co_local.strftime('%H:%M')}

    if request.method == 'POST' and request.POST.get('action') == 'save_timesheet':
        activities_json = request.POST.get('activities_data', '')
        if activities_json:
            try:
                activities_list = _clean_timesheet_activities(
                    json.loads(activities_json),
                    num_days,
                )
            except (TypeError, ValueError, json.JSONDecodeError) as exc:
                logger.warning("Invalid timesheet payload from %s: %s", request.user.username, exc)
                messages.error(request, "Timesheet data could not be saved. Please review the entries and try again.")
                return redirect(f"{reverse('timesheet')}?year={year}&month={month}")

            with transaction.atomic():
                TimesheetActivity.objects.filter(user=request.user, year=year, month=month).delete()
                for i, act_data in enumerate(activities_list):
                    TimesheetActivity.objects.create(
                        user=request.user, year=year, month=month,
                        srno=act_data.get('sr', i+1),
                        activity=act_data.get('category', ""),
                        sub_activity=act_data.get('sub', ""),
                        artifact_id=act_data.get('id', ""),
                        daily_hours=act_data.get('hours', {})
                    )
        return redirect(f"{reverse('timesheet')}?year={year}&month={month}")

    # 1. Sync/Populate Derived Timesheet Data from Monthly Roster (Source)
    att_records = AttendanceRecord.objects.filter(user=request.user, date__year=year, date__month=month)
    for r in att_records:
        if r.check_in and not hasattr(r, 'timesheet_derived'):
            # Create NEW derived record: In = Source In, Out = Source In + 9 Hours
            TimesheetRecord.objects.create(
                attendance_record=r,
                ts_check_in=r.check_in,
                ts_check_out=r.check_in + timedelta(hours=9)
            )

    # 2. Build display map using Derived data
    time_map = {r.date.day: get_billable_entry(r) for r in att_records if get_billable_entry(r)}
    db_activities = list(TimesheetActivity.objects.filter(user=request.user, year=year, month=month))
    
    if not db_activities:
        for i, (a, s) in enumerate([("Support", "Support"), ("Support", "KT"), ("Development/Analysis/Testing", "Development"), ("Development/Analysis/Testing", "Analysis"), ("Development/Analysis/Testing", "Testing")], 1):
            TimesheetActivity.objects.create(user=request.user, year=year, month=month, srno=i, activity=a, sub_activity=s)
        db_activities = list(TimesheetActivity.objects.filter(user=request.user, year=year, month=month))

    formatted_activities = []
    for a in db_activities:
        cat_class = 'cat-dev'
        low_act = a.activity.lower() if a.activity else ""
        if 'support' in low_act: cat_class = 'cat-sup'
        elif 'meet' in low_act or 'scrum' in low_act: cat_class = 'cat-meet'
        elif 'test' in low_act: cat_class = 'cat-test'
        
        formatted_activities.append({
            'sr': a.srno, 'category': a.activity or "Other", 'catClass': cat_class,
            'sub': a.sub_activity or "", 'id': a.artifact_id or "", 'hours': a.daily_hours
        })

    holiday_map = {h.date.day: h.name for h in Holiday.objects.filter(date__year=year, date__month=month)}
    for d in days:
        if is_config_holiday(date(year, month, d)): holiday_map[d] = "Holiday"

    prev_m = (current - timedelta(days=1)).replace(day=1)
    next_m = (current + timedelta(days=32)).replace(day=1)

    timesheet_context = {
        'year': year,
        'month': month,
        'time_map': time_map,
        'activities': formatted_activities,
        'holidays': holiday_map,
    }

    context = {
        'year': year, 'month': month, 'month_name': current.strftime('%B'),
        'prev_year': prev_m.year, 'prev_month': prev_m.month,
        'next_year': next_m.year, 'next_month': next_m.month,
        'timesheet_context': timesheet_context,
    }
    return render(request, 'attendance/timesheet_v2.html', context)


@login_required
def timesheet_export_view(request):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = timezone.localtime(timezone.now()).date()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if month < 1 or month > 12:
            raise ValueError
    except (TypeError, ValueError):
        year, month = today.year, today.month
    num_days = calendar.monthrange(year, month)[1]
    days = list(range(1, num_days + 1))

    att_records = AttendanceRecord.objects.filter(user=request.user, date__year=year, date__month=month)
    time_map = {r.date.day: build_billable_time_entry(r) for r in att_records if build_billable_time_entry(r)}
    activities = list(TimesheetActivity.objects.filter(user=request.user, year=year, month=month))
    db_holidays = set(Holiday.objects.filter(date__year=year, date__month=month).values_list('date__day', flat=True))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timesheet'

    header_fill = PatternFill("solid", fgColor="002060")
    date_fill = PatternFill("solid", fgColor="FFC000")
    leave_fill = PatternFill("solid", fgColor="FFFF00")
    total_fill = PatternFill("solid", fgColor="D9EAD3")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin_gray = Side(style="thin", color="B7B7B7")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    label_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    dark_label_font = Font(name="Calibri", bold=True, size=10, color="000000")
    normal_font = Font(name="Calibri", size=10, color="000000")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.freeze_panes = "F7"
    ws.sheet_view.showGridLines = False

    fixed_headers = ["Srno", "Activity", "Sub Activity", "Comments", "artfact ID/Problem id/Incident ID"]
    for col, header in enumerate(fixed_headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = label_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for d in days:
        col = 5 + d
        day_date = date(year, month, d)
        cell = ws.cell(1, col, day_date)
        cell.font = dark_label_font
        cell.fill = date_fill
        cell.alignment = center
        cell.border = border
        cell.number_format = "d-mmm"

    leaves_col = 6 + num_days
    leaves_header = ws.cell(1, leaves_col, "Leaves")
    leaves_header.font = dark_label_font
    leaves_header.fill = leave_fill
    leaves_header.alignment = center
    leaves_header.border = border

    time_labels = [("In Time", "in_time"), ("Out Time", "out_time"), ("Total Time", "total_time"), ("ESA Time", "esa_time")]
    for row_idx, (label, key) in enumerate(time_labels, 2):
        label_cell = ws.cell(row_idx, 4, label)
        label_cell.font = dark_label_font
        label_cell.alignment = center
        label_cell.border = border
        for d in days:
            col = 5 + d
            entry = time_map.get(d)
            value = entry.get(key) if entry else None
            cell = ws.cell(row_idx, col, value)
            cell.font = normal_font
            cell.alignment = center
            cell.border = border
        ws.cell(row_idx, leaves_col).border = border

    activity_start_row = 7
    for row_offset, act in enumerate(activities):
        row = activity_start_row + row_offset
        values = [
            row_offset + 1,
            act.activity,
            act.sub_activity,
            act.comments,
            act.artifact_id,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.font = normal_font
            cell.alignment = center if col in (1,) else left_wrap
            cell.fill = white_fill
            cell.border = border

        for d_str, val in act.daily_hours.items():
            try:
                day = int(d_str)
            except (TypeError, ValueError):
                continue
            if day < 1 or day > num_days:
                continue
            cell = ws.cell(row, 5 + day, val)
            cell.font = normal_font
            cell.alignment = center
            cell.border = border
        for d in days:
            ws.cell(row, 5 + d).border = border
        ws.cell(row, leaves_col).border = border

    total_row = activity_start_row + max(len(activities), 1)
    total_label = ws.cell(total_row, 5, "TOTAL")
    total_label.font = dark_label_font
    total_label.fill = total_fill
    total_label.alignment = center
    total_label.border = border
    for d in days:
        col = 5 + d
        col_letter = get_column_letter(col)
        cell = ws.cell(total_row, col, f"=SUM({col_letter}{activity_start_row}:{col_letter}{total_row - 1})")
        cell.font = dark_label_font
        cell.fill = total_fill
        cell.alignment = center
        cell.border = border
    ws.cell(total_row, leaves_col).border = border

    for row in range(1, total_row + 1):
        for col in range(1, leaves_col + 1):
            ws.cell(row, col).border = border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 24
    for col in range(6, leaves_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.row_dimensions[1].height = 52
    for row in range(activity_start_row, total_row):
        ws.row_dimensions[row].height = 36

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"Timesheet_{year}_{month:02d}.xlsx"
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
def compoff_view(request):
    if request.method == 'POST':
        worked_str = request.POST.get('worked_date', '').strip()
        reason = request.POST.get('reason', '').strip()
        try:
            worked_date = datetime.strptime(worked_str, '%Y-%m-%d').date()
            if worked_date.weekday() == 5:
                CompOffRecord.objects.get_or_create(user=request.user, worked_date=worked_date, defaults={'reason': reason})
                messages.success(request, f"Comp-off for {worked_str} logged.")
        except: pass
    return redirect('timesheet')


@login_required
def compoff_consume_view(request, compoff_id):
    if request.method == 'POST':
        leave_str = request.POST.get('leave_date', '').strip()
        try:
            record = CompOffRecord.objects.get(id=compoff_id, user=request.user)
            if leave_str:
                record.leave_date = datetime.strptime(leave_str, '%Y-%m-%d').date()
                record.status = 'consumed'
                record.save()
        except: pass
    return redirect('timesheet')


@login_required
def compoff_delete_view(request, compoff_id):
    if request.method == 'POST':
        CompOffRecord.objects.filter(id=compoff_id, user=request.user).delete()
    return redirect('timesheet')


@login_required
def support_view(request):
    if request.method == "POST":
        message_body = request.POST.get("message", "")
        subject = f"Bug Report / Support Request from {request.user.username}"
        
        try:
            send_mail(
                subject,
                f"User: {request.user.username}\nEmail: {request.user.email}\n\nMessage:\n{message_body}",
                settings.DEFAULT_FROM_EMAIL,
                [settings.EMAIL_HOST_USER],
                fail_silently=False,
            )
            messages.success(request, "Thank you! Your report has been sent to the administrator.")
        except Exception as e:
            logger.error("Failed to send support email: %s", str(e))
            messages.error(request, "Sorry, there was an error sending your report. Please try again later.")
            
        return redirect("checkin_checkout")
    return render(request, "attendance/support.html")

def privacy_policy_view(request): return render(request, 'attendance/privacy.html')
def terms_of_service_view(request): return render(request, 'attendance/terms.html')
