from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse
from django.utils import timezone
from .models import AttendanceRecord, TimesheetActivity
from datetime import date, timedelta
from io import BytesIO
import json
import calendar
from openpyxl import load_workbook

class AttendanceTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='testuser', password='password123')
        self.client = Client()
        self.client.force_login(self.user)

    def test_dashboard_load(self):
        response = self.client.get(reverse('checkin_checkout'))
        self.assertEqual(response.status_code, 200)

    def test_timesheet_load(self):
        response = self.client.get(reverse('timesheet'))
        self.assertEqual(response.status_code, 200)

    def test_invalid_timesheet_params(self):
        # Test that bad year/month don't crash the server
        response = self.client.get(reverse('timesheet'), {'year': 'abc', 'month': '99'})
        self.assertEqual(response.status_code, 200)

    def test_checkin_action_records_time(self):
        response = self.client.post(reverse('checkin_checkout'), {'action': 'check_in'})
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            AttendanceRecord.objects.filter(user=self.user, date=date.today(), check_in__isnull=False).exists()
        )

    def test_timesheet_save_atomic(self):
        # Simple save test through the current JSON payload used by the UI
        post_data = {
            'action': 'save_timesheet',
            'activities_data': json.dumps([
                {
                    'sr': 1,
                    'category': 'Test Task',
                    'sub': 'Implementation',
                    'id': 'TSK-1',
                    'hours': {'1': '8'},
                }
            ]),
        }
        response = self.client.post(reverse('timesheet'), post_data)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(TimesheetActivity.objects.filter(user=self.user).count(), 1)

    def test_timesheet_payload_rejects_invalid_hours(self):
        response = self.client.post(reverse('timesheet'), {
            'action': 'save_timesheet',
            'activities_data': json.dumps([
                {'sr': 1, 'category': 'Bad', 'sub': '', 'id': '', 'hours': {'1': '25'}},
            ]),
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(TimesheetActivity.objects.filter(user=self.user).count(), 0)

    def test_timesheet_json_script_escapes_activity_text(self):
        TimesheetActivity.objects.create(
            user=self.user,
            year=date.today().year,
            month=date.today().month,
            srno=1,
            activity='</script><img src=x onerror=alert(1)>',
            sub_activity='Review',
            artifact_id='A-1',
            daily_hours={'1': 1},
        )
        response = self.client.get(reverse('timesheet'))
        self.assertEqual(response.status_code, 200)
        self.assertNotIn(b'</script><img src=x onerror=alert(1)>', response.content)

    def test_month_excel_export_handles_timezone_datetimes(self):
        now = timezone.now()
        AttendanceRecord.objects.create(
            user=self.user,
            date=now.date(),
            check_in=now,
            check_out=now + timedelta(hours=8),
        )
        response = self.client.get(reverse('month_excel_export'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_timesheet_export_uses_nse_grid_headers(self):
        TimesheetActivity.objects.create(
            user=self.user,
            year=date.today().year,
            month=date.today().month,
            srno=1,
            activity='Development',
            sub_activity='Enhancement',
            artifact_id='ART-1',
            daily_hours={'1': 4},
        )
        response = self.client.get(reverse('timesheet_export'))
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content), data_only=False)
        ws = wb.active
        self.assertEqual(ws['A1'].value, 'Srno')
        self.assertEqual(ws['B1'].value, 'Activity')
        self.assertEqual(ws['C1'].value, 'Sub Activity')
        self.assertEqual(ws['D1'].value, 'Comments')
        self.assertEqual(ws['E1'].value, 'artfact ID/Problem id/Incident ID')
        self.assertEqual(ws['F1'].value.day, 1)
        _, days_in_month = calendar.monthrange(date.today().year, date.today().month)
        self.assertEqual(ws.cell(1, 6 + days_in_month).value, 'Leaves')
        self.assertEqual(ws['B7'].value, 'Development')
        self.assertEqual(ws['E7'].value, 'ART-1')
